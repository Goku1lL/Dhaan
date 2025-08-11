"""Excel integration and management for the Algo Trading System."""

import pandas as pd
from pathlib import Path
from typing import Optional, List, Dict, Any
import logging

# Configure logging
logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. Excel features will be disabled.")


class ExcelManager:
    """Manages Excel workbook connections and operations."""
    
    def __init__(self, live_trade_file: Path, completed_orders_file: Path):
        """
        Initialize Excel manager.
        
        Args:
            live_trade_file: Path to live trade data Excel file
            completed_orders_file: Path to completed orders Excel file
        """
        self.live_trade_file = live_trade_file
        self.completed_orders_file = completed_orders_file
        
        # Excel objects
        self.live_wb = None
        self.live_trading_sheet = None
        self.completed_wb = None
        self.completed_order_sheet = None
        
        if OPENPYXL_AVAILABLE:
            self._connect_to_workbooks()
        else:
            logger.error("openpyxl not available. Excel operations disabled.")
    
    def _connect_to_workbooks(self):
        """Connect to Excel workbooks and sheets."""
        try:
            # Load live trade data workbook
            if self.live_trade_file.exists():
                self.live_wb = load_workbook(str(self.live_trade_file))
                self.live_trading_sheet = self.live_wb.active
                logger.info(f"Connected to live trade workbook: {self.live_trade_file}")
            else:
                # Create new workbook if it doesn't exist
                self.live_wb = Workbook()
                self.live_trading_sheet = self.live_wb.active
                self.live_trading_sheet.title = "Live Trading"
                self.live_wb.save(str(self.live_trade_file))
                logger.info(f"Created new live trade workbook: {self.live_trade_file}")
            
            # Load completed orders workbook
            if self.completed_orders_file.exists():
                self.completed_wb = load_workbook(str(self.completed_orders_file))
                self.completed_order_sheet = self.completed_wb.active
                logger.info(f"Connected to completed orders workbook: {self.completed_orders_file}")
            else:
                # Create new workbook if it doesn't exist
                self.completed_wb = Workbook()
                self.completed_order_sheet = self.completed_wb.active
                self.completed_order_sheet.title = "Completed Orders"
                self.completed_wb.save(str(self.completed_orders_file))
                logger.info(f"Created new completed orders workbook: {self.completed_orders_file}")
            
        except Exception as e:
            logger.error(f"Failed to connect to Excel workbooks: {e}")
            self.live_wb = self.live_trading_sheet = None
            self.completed_wb = self.completed_order_sheet = None
    
    def clear_sheets(self, clear_range: str = "A2:Z100"):
        """
        Clear existing data from both sheets.
        
        Args:
            clear_range: Range to clear (default: A2:Z100)
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("Cannot clear sheets: openpyxl not available")
            return
        
        try:
            # Clear live trading sheet
            if self.live_trading_sheet:
                # Clear cells from A2 to Z100
                for row in range(2, 101):
                    for col in range(1, 27):  # A to Z
                        cell = self.live_trading_sheet.cell(row=row, column=col)
                        cell.value = None
                logger.info("Cleared live trading sheet")
            
            # Clear completed orders sheet
            if self.completed_order_sheet:
                # Clear cells from A2 to Z100
                for row in range(2, 101):
                    for col in range(1, 27):  # A to Z
                        cell = self.completed_order_sheet.cell(row=row, column=col)
                        cell.value = None
                logger.info("Cleared completed orders sheet")
                
        except Exception as e:
            logger.error(f"Failed to clear sheets: {e}")
    
    def read_watchlist(self) -> List[str]:
        """
        Read watchlist from Excel starting from A2.
        
        Returns:
            List of symbol names
        """
        if not OPENPYXL_AVAILABLE or not self.live_trading_sheet:
            logger.warning("Cannot read watchlist: openpyxl or sheet not available")
            return []
        
        try:
            watchlist = []
            row = 2
            while True:
                cell_value = self.live_trading_sheet.cell(row=row, column=1).value
                if cell_value is None:
                    break
                watchlist.append(str(cell_value).strip())
                row += 1
            
            logger.info(f"Read watchlist: {watchlist}")
            return watchlist
            
        except Exception as e:
            logger.error(f"Failed to read watchlist: {e}")
            return []
    
    def read_quantity(self, symbol: str, full_watchlist: List[str]) -> int:
        """
        Read quantity for a specific symbol from Excel.
        
        Args:
            symbol: Symbol name
            full_watchlist: Full watchlist to find row number
            
        Returns:
            Quantity for the symbol
        """
        if not OPENPYXL_AVAILABLE or not self.live_trading_sheet:
            return 1  # Default quantity
        
        try:
            row_number = full_watchlist.index(symbol) + 2
            qty_cell = self.live_trading_sheet.cell(row=row_number, column=7).value  # Column G
            if qty_cell:
                return int(qty_cell)
        except Exception as e:
            logger.warning(f"Failed to read quantity for {symbol}: {e}")
        
        return 1  # Default quantity
    
    def read_reentry_flag(self) -> str:
        """
        Read re-entry flag from Excel cell M1.
        
        Returns:
            Re-entry flag value
        """
        if not OPENPYXL_AVAILABLE or not self.live_trading_sheet:
            return "No"
        
        try:
            re_entry = self.live_trading_sheet.cell(row=1, column=13).value  # Column M
            return str(re_entry) if re_entry else "No"
        except Exception as e:
            logger.warning(f"Failed to read re-entry flag: {e}")
            return "No"
    
    def update_live_trading_sheet(self, order_book: Dict[str, Any]):
        """
        Update live trading sheet with current order book.
        
        Args:
            order_book: Dictionary of order data
        """
        if not OPENPYXL_AVAILABLE or not self.live_trading_sheet:
            logger.warning("Cannot update live trading sheet: openpyxl or sheet not available")
            return
        
        try:
            if order_book:
                df_live = pd.DataFrame(order_book).T
                # Write DataFrame to Excel sheet
                for i, (symbol, data) in enumerate(df_live.iterrows()):
                    for j, (col, value) in enumerate(data.items()):
                        if pd.notna(value):
                            self.live_trading_sheet.cell(row=i+1, column=j+1, value=value)
                logger.debug("Updated live trading sheet")
        except Exception as e:
            logger.error(f"Failed to update live trading sheet: {e}")
    
    def update_completed_orders_sheet(self, completed_orders: List[Any]):
        """
        Update completed orders sheet with completed trades.
        
        Args:
            completed_orders: List of completed order data
        """
        if not OPENPYXL_AVAILABLE or not self.completed_order_sheet:
            logger.warning("Cannot update completed orders sheet: openpyxl or sheet not available")
            return
        
        try:
            if completed_orders:
                df_completed = pd.DataFrame(completed_orders)
                # Write DataFrame to Excel sheet
                for i, row_data in enumerate(df_completed.iterrows()):
                    for j, (col, value) in enumerate(row_data[1].items()):
                        if pd.notna(value):
                            self.completed_order_sheet.cell(row=i+1, column=j+1, value=value)
                logger.debug("Updated completed orders sheet")
        except Exception as e:
            logger.error(f"Failed to update completed orders sheet: {e}")
    
    def close_workbooks(self):
        """Close all Excel workbooks."""
        if not OPENPYXL_AVAILABLE:
            return
        
        try:
            if self.live_wb:
                self.live_wb.save(str(self.live_trade_file))
                logger.info("Saved live trade workbook")
            
            if self.completed_wb:
                self.completed_wb.save(str(self.completed_orders_file))
                logger.info("Saved completed orders workbook")
                
        except Exception as e:
            logger.error(f"Failed to save workbooks: {e}")
    
    def is_available(self) -> bool:
        """Check if Excel functionality is available."""
        return OPENPYXL_AVAILABLE and self.live_trading_sheet is not None 